import os
import streamlit as st
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load environment variables (if using .env file)
load_dotenv()

# Fetch API URL from environment variable
api_url = os.environ.get("API_URL")

# Title and Description
col1, col2, col3 = st.columns([1.5, 5, 1.5])  # Define relative column width (e.g., 1 for image, 5 for title)

with col1:
    st.image("Safety.png", width=100)  # Image on the left

with col2:
    st.title("MB - Health Insurance Premium Calculator")  # Title on the right

with col3:
    st.image("Safety.png", width=100)  # Image on the left
st.markdown("<h6 style='text-align: center;'>Estimate your insurance premium based on your personal and health data.</h6>", unsafe_allow_html=True)

st.markdown("""
    <style>
        .form-container {
            border-radius: 50px;
        }
    </style>
""", unsafe_allow_html=True)

# Create a container for the form
with st.container():
    with st.form(key="insurance_form", clear_on_submit=False):
        # Centered Heading for User Details Section
        st.markdown("<h3 style='text-align: center;'>Add your details to get Predicted Premium</h3>", unsafe_allow_html=True)

        # Create two columns side by side for Age and BMI
        col1, col2 = st.columns(2)
        with col1:
            age = st.number_input("Age", min_value=18, max_value=66, step=1, help="Enter your age (18-66).")
        with col2:
            bmi = st.number_input("BMI", min_value=10.0, max_value=50.0, step=0.1, help="Enter your Body Mass Index (BMI).")

        # Create another two columns for Weight and Height
        col3, col4 = st.columns(2)
        with col3:
            weight = st.number_input("Weight (in kg)", min_value=30, max_value=200, step=1, help="Enter your weight in kilograms.")
        with col4:
            height = st.number_input("Height (in cm)", min_value=100, max_value=250, step=1, help="Enter your height in centimeters.")

        # Create another two columns for Diabetes and Blood Pressure
        col5, col6 = st.columns(2)
        with col5:
            diabetes = st.radio("Do you have diabetes?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if you have diabetes.")
        with col6:
            blood_pressure = st.radio("Do you have blood pressure problems?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if you have blood pressure issues.")

        # Create another two columns for Transplants and Chronic Diseases
        col7, col8 = st.columns(2)
        with col7:
            transplants = st.radio("Have you undergone any transplants?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if you've had any transplants.")
        with col8:
            chronic_diseases = st.radio("Do you have any chronic diseases?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if you have chronic illnesses.")

        # Create another two columns for Allergies and Cancer History
        col9, col10 = st.columns(2)
        with col9:
            allergies = st.radio("Do you have known allergies?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if you have allergies.")
        with col10:
            cancer_history = st.radio("Is there a history of cancer in your family?", options=[0, 1], format_func=lambda x: "Yes" if x else "No", help="Select 'Yes' if there's a family history of cancer.")

        # Create one more field for Major Surgeries
        major_surgeries = st.selectbox("Number of major surgeries", options=[0, 1, 2, 3], help="Enter the number of major surgeries you've had.")

        # Condition to check if all required fields are filled
        all_filled = all([age, bmi, weight, height, diabetes is not None, blood_pressure is not None, transplants is not None,
                          chronic_diseases is not None, allergies is not None, cancer_history is not None, major_surgeries is not None])

        # Submit Button for Prediction (disabled until all fields are filled)
        submit_button = st.form_submit_button(label="Calculate Premium", disabled=not all_filled)

        if submit_button:
            # Prepare input data for API
            input_data = {
                "Age": age,
                "Diabetes": diabetes,
                "BloodPressureProblems": blood_pressure,
                "AnyTransplants": transplants,
                "AnyChronicDiseases": chronic_diseases,
                "Height": height,
                "Weight": weight,
                "KnownAllergies": allergies,
                "HistoryOfCancerInFamily": cancer_history,
                "NumberOfMajorSurgeries": major_surgeries,
                "BMI": bmi
            }

            # Send input data to Flask API
            try:
                response = requests.post(api_url, json=input_data)
                if response.status_code == 200:
                    # Extract prediction
                    premium = response.json().get("PredictedPremium")

                    # Show notification with the prediction
                    st.success(f"Predicted Insurance Premium: {premium}", icon="💸")

                    # Create DataFrame from input data
                    input_data["PredictedPremium"] = premium  # Add premium to data

                    # Define the Excel file path
                    file_path = "insurance_data.xlsx"

                    # Check if the Excel file exists
                    try:
                        # Try to open the existing workbook
                        wb = load_workbook(file_path)
                        sheet = wb.active
                    except FileNotFoundError:
                        # If the file doesn't exist, create a new one with headers
                        wb = load_workbook(file_path) if os.path.exists(file_path) else None
                        if not wb:
                            wb = Workbook()  # Create a new workbook if it doesn't exist
                        sheet = wb.active
                        sheet.append(list(input_data.keys()))  # Add headers
                        wb.save(file_path)

                    # Append the user input data to the sheet
                    sheet.append(list(input_data.values()))
                    wb.save(file_path)

                else:
                    st.error("Failed to retrieve prediction. Please check the API.", icon="❌")
            except Exception as e:
                st.error(f"Error connecting to the backend: {e}")

        st.markdown('</div>', unsafe_allow_html=True)
